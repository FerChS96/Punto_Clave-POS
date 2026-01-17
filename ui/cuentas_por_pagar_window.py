"""
Ventana de Cuentas por Pagar para HTF POS
Usando componentes reutilizables del sistema de diseño
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem,
    QHeaderView, QDateEdit, QSizePolicy, QComboBox, QAbstractItemView, QDialog, QLabel
)
from PySide6.QtCore import Qt, Signal, QDate, QTimer
from PySide6.QtGui import QFont
import logging
from datetime import datetime
import qtawesome as qta

# Importar componentes del sistema de diseño
from ui.components import (
    WindowsPhoneTheme,
    TileButton,
    CompactNavButton,
    create_page_layout,
    ContentPanel,
    StyledLabel,
    SearchBar,
    show_info_dialog,
    show_warning_dialog,
    show_error_dialog,
    aplicar_estilo_fecha
)


class CuentasPorPagarWindow(QWidget):
    """Widget para ver cuentas por pagar"""

    cerrar_solicitado = Signal()

    def __init__(self, pg_manager, user_data, parent=None):
        super().__init__(parent)
        self.pg_manager = pg_manager
        self.user_data = user_data
        self.cuentas_data = []  # Almacenar todas las cuentas cargadas
        self.cuentas_filtradas = []  # Cuentas después de aplicar filtros
        self.pagina_actual = 0
        self.items_por_pagina = 50

        # Timer para detectar entrada del escáner
        self.scanner_timer = QTimer()
        self.scanner_timer.setSingleShot(True)
        self.scanner_timer.setInterval(300)  # 300ms después de que deje de escribir
        self.scanner_timer.timeout.connect(self.aplicar_filtros)

        # Configurar política de tamaño
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.setup_ui()

    def setup_ui(self):
        """Configurar interfaz de cuentas por pagar"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Contenido
        content = QWidget()
        content_layout = create_page_layout("")
        content.setLayout(content_layout)
        
        # Buscador
        self.search_bar = SearchBar("Buscar por número de cuenta, proveedor o monto...")
        self.search_bar.connect_search(self.on_search_changed)
        content_layout.addWidget(self.search_bar)
        
        # Filtros
        self.create_filters(content_layout)
        
        # Tabla
        self.create_history_table(content_layout)
        
        # Panel de información y paginación
        self.create_info_buttons_panel(content_layout)
        
        # Botones
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(WindowsPhoneTheme.TILE_SPACING)
        
        btn_exportar = TileButton("Exportar", "fa5s.download", WindowsPhoneTheme.TILE_GREEN)
        btn_exportar.clicked.connect(self.exportar_datos)
        
        btn_cerrar = TileButton("Cerrar", "fa5s.times", WindowsPhoneTheme.TILE_RED)
        btn_cerrar.clicked.connect(self.cerrar_solicitado.emit)
        
        buttons_layout.addWidget(btn_exportar)
        buttons_layout.addWidget(btn_cerrar)
        
        content_layout.addLayout(buttons_layout)
        layout.addWidget(content)
        
        # Cargar datos iniciales
        self.cargar_cuentas_completo()

    def create_filters(self, parent_layout):
        """Crear filtros de búsqueda"""
        filters_panel = ContentPanel()
        filters_layout = QHBoxLayout(filters_panel)
        filters_layout.setSpacing(WindowsPhoneTheme.MARGIN_MEDIUM)
        
        # Fecha desde
        desde_container = QWidget()
        desde_layout = QVBoxLayout(desde_container)
        desde_layout.setContentsMargins(0, 0, 0, 0)
        desde_layout.setSpacing(4)
        desde_label = StyledLabel("Desde:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        desde_layout.addWidget(desde_label)
        self.fecha_desde = QDateEdit()
        self.fecha_desde.setDate(QDate.currentDate().addDays(-30))
        self.fecha_desde.setCalendarPopup(True)
        self.fecha_desde.setMinimumHeight(40)
        self.fecha_desde.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        self.fecha_desde.dateChanged.connect(self.cargar_cuentas_completo)
        aplicar_estilo_fecha(self.fecha_desde)
        desde_layout.addWidget(self.fecha_desde)
        filters_layout.addWidget(desde_container, stretch=1)
        
        # Fecha hasta
        hasta_container = QWidget()
        hasta_layout = QVBoxLayout(hasta_container)
        hasta_layout.setContentsMargins(0, 0, 0, 0)
        hasta_layout.setSpacing(4)
        hasta_label = StyledLabel("Hasta:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        hasta_layout.addWidget(hasta_label)
        self.fecha_hasta = QDateEdit()
        self.fecha_hasta.setDate(QDate.currentDate())
        self.fecha_hasta.setCalendarPopup(True)
        self.fecha_hasta.setMinimumHeight(40)
        self.fecha_hasta.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        self.fecha_hasta.dateChanged.connect(self.cargar_cuentas_completo)
        aplicar_estilo_fecha(self.fecha_hasta)
        hasta_layout.addWidget(self.fecha_hasta)
        filters_layout.addWidget(hasta_container, stretch=1)
        
        # Filtro por estado
        estado_container = QWidget()
        estado_layout = QVBoxLayout(estado_container)
        estado_layout.setContentsMargins(0, 0, 0, 0)
        estado_layout.setSpacing(4)
        estado_label = StyledLabel("Estado:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        estado_layout.addWidget(estado_label)
        self.estado_combo = QComboBox()
        self.estado_combo.setMinimumHeight(40)
        self.estado_combo.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        self.estado_combo.addItems(["Todos", "activa", "pagada", "vencida", "cancelada"])
        self.estado_combo.setCurrentIndex(0)
        self.estado_combo.currentTextChanged.connect(self.aplicar_filtros)
        estado_layout.addWidget(self.estado_combo)
        filters_layout.addWidget(estado_container, stretch=1)
        
        # Botón limpiar filtros
        btn_limpiar_container = QWidget()
        btn_limpiar_layout = QVBoxLayout(btn_limpiar_container)
        btn_limpiar_layout.setContentsMargins(0, 0, 0, 0)
        btn_limpiar_layout.setSpacing(4)
        limpiar_spacer = StyledLabel("", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        btn_limpiar_layout.addWidget(limpiar_spacer)
        btn_limpiar = QPushButton("Limpiar")
        btn_limpiar.setMinimumHeight(40)
        btn_limpiar.setMinimumWidth(100)
        btn_limpiar.setObjectName("tileButton")
        btn_limpiar.setProperty("tileColor", WindowsPhoneTheme.TILE_ORANGE)
        btn_limpiar.clicked.connect(self.limpiar_filtros)
        btn_limpiar_layout.addWidget(btn_limpiar)
        filters_layout.addWidget(btn_limpiar_container)
        
        parent_layout.addWidget(filters_panel)

    def create_history_table(self, parent_layout):
        """Crear tabla de cuentas por pagar"""
        self.history_table = QTableWidget()
        self.history_table.setColumnCount(7)
        self.history_table.setHorizontalHeaderLabels([
            "Número", "Fecha", "Proveedor", "Total", "Saldo", "Estado", "Detalles"
        ])
        
        # Configurar para que no sea editable
        self.history_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.history_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.history_table.setSelectionMode(QAbstractItemView.SingleSelection)
        
        header = self.history_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)
        
        parent_layout.addWidget(self.history_table)

    def create_info_buttons_panel(self, parent_layout):
        """Crear panel de información con paginación integrada"""
        info_panel = ContentPanel()
        info_layout = QHBoxLayout(info_panel)
        info_layout.setSpacing(8)

        # Etiqueta de información (con paginación integrada)
        self.info_label = StyledLabel("", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        info_layout.addWidget(self.info_label, stretch=1)

        # Botones de paginación
        self.btn_pagina_anterior = CompactNavButton(
            "Anterior",
            "fa5s.chevron-left",
            WindowsPhoneTheme.TILE_BLUE,
            icon_position="left"
        )
        self.btn_pagina_anterior.setToolTip("Página anterior")
        self.btn_pagina_anterior.clicked.connect(self.pagina_anterior)
        info_layout.addWidget(self.btn_pagina_anterior)

        self.btn_proxima_pagina = CompactNavButton(
            "Siguiente",
            "fa5s.chevron-right",
            WindowsPhoneTheme.TILE_BLUE,
            icon_position="right"
        )
        self.btn_proxima_pagina.setToolTip("Página siguiente")
        self.btn_proxima_pagina.clicked.connect(self.proxima_pagina)
        info_layout.addWidget(self.btn_proxima_pagina)

        parent_layout.addWidget(info_panel)

    def on_search_changed(self, text):
        """Reiniciar timer cuando cambia el texto de búsqueda"""
        self.scanner_timer.start()

    def cargar_cuentas_completo(self):
        """Cargar cuentas por pagar desde la base de datos"""
        try:
            fecha_desde = self.fecha_desde.date().toPython()
            fecha_hasta = self.fecha_hasta.date().toPython()

            # Consulta PostgreSQL
            cuentas = self.pg_manager.query("""
                SELECT
                    cxp.id_cuenta_pagar,
                    cxp.numero_cuenta,
                    cxp.fecha_cuenta,
                    prov.razon_social as proveedor,
                    cxp.total,
                    cxp.saldo,
                    cxp.estado
                FROM cuentas_por_pagar cxp
                LEFT JOIN ca_proveedores prov ON cxp.id_proveedor = prov.id_proveedor
                WHERE cxp.fecha_cuenta >= %s AND cxp.fecha_cuenta <= %s
                ORDER BY cxp.fecha_cuenta DESC
            """, (fecha_desde, fecha_hasta))

            # Convertir a formato esperado
            self.cuentas_data = []
            for cuenta in cuentas:
                cuenta_dict = {
                    'id_cuenta_pagar': cuenta['id_cuenta_pagar'],
                    'numero_cuenta': cuenta['numero_cuenta'],
                    'fecha_cuenta': cuenta['fecha_cuenta'],
                    'proveedor': cuenta['proveedor'] or 'N/A',
                    'total': cuenta['total'],
                    'saldo': cuenta['saldo'],
                    'estado': cuenta['estado']
                }
                self.cuentas_data.append(cuenta_dict)

            # Aplicar filtros
            self.aplicar_filtros()

        except Exception as e:
            logging.error(f"Error cargando cuentas por pagar: {e}")
            show_warning_dialog(self, "Error", f"Error al cargar cuentas: {e}")

    def aplicar_filtros(self):
        """Aplicar filtros a los datos de cuentas"""
        try:
            # Filtro de búsqueda
            search_text = self.search_bar.text().lower().strip()

            # Filtro de estado
            estado_filtro = self.estado_combo.currentText()
            if estado_filtro == "Todos":
                estado_filtro = None

            # Aplicar filtros
            self.cuentas_filtradas = []
            for cuenta in self.cuentas_data:
                # Filtro de búsqueda
                if search_text:
                    searchable = f"{cuenta['numero_cuenta']} {cuenta['proveedor']} {cuenta['total']}".lower()
                    if search_text not in searchable:
                        continue

                # Filtro de estado
                if estado_filtro and cuenta['estado'] != estado_filtro:
                    continue

                self.cuentas_filtradas.append(cuenta)

            # Actualizar paginación
            self.pagina_actual = 0
            self.actualizar_tabla()

        except Exception as e:
            logging.error(f"Error aplicando filtros: {e}")

    def actualizar_tabla(self):
        """Actualizar tabla con datos filtrados y paginados"""
        try:
            # Calcular índices de paginación
            inicio = self.pagina_actual * self.items_por_pagina
            fin = inicio + self.items_por_pagina
            cuentas_pagina = self.cuentas_filtradas[inicio:fin]

            # Configurar tabla
            self.history_table.setRowCount(len(cuentas_pagina))

            for row, cuenta in enumerate(cuentas_pagina):
                # Número de cuenta
                self.history_table.setItem(row, 0, QTableWidgetItem(cuenta['numero_cuenta']))

                # Fecha
                fecha_str = cuenta['fecha_cuenta'].strftime('%d/%m/%Y') if cuenta['fecha_cuenta'] else 'N/A'
                self.history_table.setItem(row, 1, QTableWidgetItem(fecha_str))

                # Proveedor
                self.history_table.setItem(row, 2, QTableWidgetItem(cuenta['proveedor']))

                # Total
                self.history_table.setItem(row, 3, QTableWidgetItem(f"${cuenta['total']:.2f}"))

                # Saldo
                self.history_table.setItem(row, 4, QTableWidgetItem(f"${cuenta['saldo']:.2f}"))

                # Estado
                self.history_table.setItem(row, 5, QTableWidgetItem(cuenta['estado'].title()))

                # Botón detalles con icono
                btn_detalles = QPushButton()
                btn_detalles.setIcon(qta.icon('fa5s.eye', color='white'))
                btn_detalles.setToolTip("Ver detalle de la cuenta")
                btn_detalles.setFixedWidth(40)
                btn_detalles.setMinimumHeight(35)
                btn_detalles.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {WindowsPhoneTheme.TILE_BLUE};
                        color: white;
                        border: none;
                        border-radius: 3px;
                    }}
                    QPushButton:hover {{
                        background-color: #1976d2;
                    }}
                """)
                btn_detalles.clicked.connect(lambda checked, cid=cuenta['id_cuenta_pagar']: self.ver_detalles_cuenta(cid))
                self.history_table.setCellWidget(row, 6, btn_detalles)

            # Actualizar controles de paginación
            total_paginas = (len(self.cuentas_filtradas) + self.items_por_pagina - 1) // self.items_por_pagina
            self.btn_pagina_anterior.setEnabled(self.pagina_actual > 0)
            self.btn_proxima_pagina.setEnabled(self.pagina_actual < total_paginas - 1)

            # Actualizar etiqueta de información
            total_cuentas = len(self.cuentas_filtradas)
            if total_cuentas > 0:
                self.info_label.setText(
                    f"Mostrando {inicio + 1}-{min(fin, total_cuentas)} de {total_cuentas} cuentas "
                    f"(Página {self.pagina_actual + 1} de {total_paginas})"
                )
            else:
                self.info_label.setText("No se encontraron cuentas por pagar")

        except Exception as e:
            logging.error(f"Error actualizando tabla: {e}")

    def pagina_anterior(self):
        """Ir a página anterior"""
        if self.pagina_actual > 0:
            self.pagina_actual -= 1
            self.actualizar_tabla()

    def proxima_pagina(self):
        """Ir a página siguiente"""
        total_paginas = (len(self.cuentas_filtradas) + self.items_por_pagina - 1) // self.items_por_pagina
        if self.pagina_actual < total_paginas - 1:
            self.pagina_actual += 1
            self.actualizar_tabla()

    def limpiar_filtros(self):
        """Limpiar todos los filtros"""
        self.search_bar.clear()
        self.estado_combo.setCurrentIndex(0)
        self.fecha_desde.setDate(QDate.currentDate().addDays(-30))
        self.fecha_hasta.setDate(QDate.currentDate())
        self.cargar_cuentas_completo()

    def ver_detalles_cuenta(self, id_cuenta):
        """Ver detalles de una cuenta por pagar"""
        try:
            # Obtener detalles de la cuenta
            cuenta = self.pg_manager.query("""
                SELECT cxp.*, prov.razon_social, tcp.nombre as tipo_cuenta
                FROM cuentas_por_pagar cxp
                LEFT JOIN ca_proveedores prov ON cxp.id_proveedor = prov.id_proveedor
                LEFT JOIN ca_tipo_cuenta_pagar tcp ON cxp.id_tipo_cuenta_pagar = tcp.id_tipo_cuenta_pagar
                WHERE cxp.id_cuenta_pagar = %s
            """, (id_cuenta,))

            if not cuenta:
                show_warning_dialog(self, "Error", "No se encontró la cuenta por pagar")
                return

            cuenta = cuenta[0]

            # Obtener detalles de productos si existen
            detalles = self.pg_manager.query("""
                SELECT dp.*, p.nombre
                FROM cxp_detalle_productos dp
                LEFT JOIN ca_productos p ON dp.id_producto = p.id_producto
                WHERE dp.id_cuenta_pagar = %s
            """, (id_cuenta,))

            # Crear diálogo de detalles
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Detalles - {cuenta['numero_cuenta']}")
            dialog.setModal(True)
            dialog.resize(600, 400)

            layout = QVBoxLayout(dialog)

            # Información general
            info_layout = QVBoxLayout()

            info_layout.addWidget(StyledLabel(f"Número: {cuenta['numero_cuenta']}", bold=True))
            info_layout.addWidget(StyledLabel(f"Tipo: {cuenta['tipo_cuenta'] or 'N/A'}"))
            info_layout.addWidget(StyledLabel(f"Proveedor: {cuenta['razon_social'] or 'N/A'}"))
            info_layout.addWidget(StyledLabel(f"Fecha: {cuenta['fecha_cuenta'].strftime('%d/%m/%Y')}"))
            info_layout.addWidget(StyledLabel(f"Total: ${cuenta['total']:.2f}"))
            info_layout.addWidget(StyledLabel(f"Saldo: ${cuenta['saldo']:.2f}"))
            info_layout.addWidget(StyledLabel(f"Estado: {cuenta['estado'].title()}"))

            if cuenta['numero_factura']:
                info_layout.addWidget(StyledLabel(f"Factura: {cuenta['numero_factura']}"))

            layout.addLayout(info_layout)

            # Detalles de productos si existen
            if detalles:
                layout.addWidget(StyledLabel("Productos:", bold=True))
                productos_text = ""
                for detalle in detalles:
                    productos_text += f"- {detalle['nombre'] or 'N/A'}: {detalle['cantidad']} x ${detalle['precio_unitario']:.2f} = ${detalle['subtotal_linea']:.2f}\n"
                productos_label = QLabel(productos_text)
                productos_label.setWordWrap(True)
                layout.addWidget(productos_label)

            # Botón cerrar
            btn_cerrar = TileButton("Cerrar", "fa5s.times", WindowsPhoneTheme.TILE_RED)
            btn_cerrar.clicked.connect(dialog.accept)
            layout.addWidget(btn_cerrar)

            dialog.exec()

        except Exception as e:
            logging.error(f"Error mostrando detalles: {e}")
            show_error_dialog(self, "Error", f"Error al mostrar detalles: {e}")

    def exportar_datos(self):
        """Exportar datos a archivo Excel"""
        try:
            from datetime import datetime
            import os
            
            # Verificar si openpyxl está disponible
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                show_warning_dialog(
                    self,
                    "Biblioteca requerida",
                    "Para exportar a Excel necesitas instalar openpyxl",
                    detail="Ejecuta: pip install openpyxl"
                )
                return
            
            # Obtener rango de fechas
            fecha_desde = self.fecha_desde.date().toPython()
            fecha_hasta = self.fecha_hasta.date().toPython()
            
            # Obtener datos de cuentas por pagar usando PostgreSQL
            cuentas_raw = self.pg_manager.query("""
                SELECT
                    cxp.numero_cuenta,
                    cxp.fecha_cuenta,
                    prov.razon_social as proveedor,
                    cxp.total,
                    cxp.saldo,
                    cxp.estado,
                    cxp.numero_factura
                FROM cuentas_por_pagar cxp
                LEFT JOIN ca_proveedores prov ON cxp.id_proveedor = prov.id_proveedor
                WHERE cxp.fecha_cuenta >= %s AND cxp.fecha_cuenta <= %s
                ORDER BY cxp.fecha_cuenta DESC
            """, (fecha_desde, fecha_hasta))
            
            # Crear libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Cuentas por Pagar"
            
            # Estilos
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Encabezados
            headers = ["Número Cuenta", "Fecha", "Proveedor", "Total", "Saldo", "Estado", "Factura"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Datos
            for row, cuenta in enumerate(cuentas_raw, 2):
                ws.cell(row=row, column=1, value=cuenta['numero_cuenta']).border = border
                
                # Fecha
                fecha = cuenta['fecha_cuenta'].strftime("%d/%m/%Y") if isinstance(cuenta['fecha_cuenta'], datetime) else str(cuenta['fecha_cuenta'])
                ws.cell(row=row, column=2, value=fecha).border = border
                
                # Proveedor
                ws.cell(row=row, column=3, value=cuenta['proveedor'] or 'N/A').border = border
                
                # Total
                total_cell = ws.cell(row=row, column=4, value=cuenta['total'])
                total_cell.number_format = '$#,##0.00'
                total_cell.border = border
                
                # Saldo
                saldo_cell = ws.cell(row=row, column=5, value=cuenta['saldo'])
                saldo_cell.number_format = '$#,##0.00'
                saldo_cell.border = border
                
                # Estado
                ws.cell(row=row, column=6, value=cuenta['estado'].title()).border = border
                
                # Factura
                ws.cell(row=row, column=7, value=cuenta['numero_factura'] or 'N/A').border = border
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 15
            
            # Guardar archivo
            fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            filename = os.path.join(desktop, f"cuentas_por_pagar_{fecha_str}.xlsx")
            
            wb.save(filename)
            
            show_info_dialog(
                self,
                "Exportación completada",
                f"Las cuentas por pagar han sido exportadas exitosamente",
                detail=f"Archivo guardado en:\n{filename}"
            )
            
            logging.info(f"Cuentas por pagar exportadas: {filename}")
            
        except Exception as e:
            logging.error(f"Error exportando datos: {e}")
            show_warning_dialog(
                self,
                "Error al exportar",
                "No se pudo exportar las cuentas por pagar",
                detail=str(e)
            )
"""
Ventana de Historial de Movimientos de Inventario para HTF POS
Usando componentes reutilizables del sistema de diseño
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QTableWidget, QTableWidgetItem,
    QHeaderView, QLineEdit, QSizePolicy, QFrame,
    QComboBox, QDateEdit, QLabel
)
from PySide6.QtCore import Qt, Signal, QDate, QThread, QTimer
from PySide6.QtGui import QFont
from datetime import datetime, timedelta
import logging

# Importar componentes del sistema de diseño
from ui.components import (
    WindowsPhoneTheme,
    TileButton,
    CompactNavButton,
    create_page_layout,
    ContentPanel,
    StyledLabel,
    SearchBar,
    show_info_dialog,
    show_warning_dialog,
    show_error_dialog,
    aplicar_estilo_fecha
)


class MovimientosLoaderThread(QThread):
    """Hilo para cargar movimientos de forma asíncrona"""
    
    movimientos_loaded = Signal(list)
    error_occurred = Signal(str)
    
    def __init__(self, pg_manager, limite=50):
        super().__init__()
        self.pg_manager = pg_manager
        self.limite = limite
        self._is_running = True
        # Marcar thread como daemon para que no bloquee la aplicación
        self.setTerminationEnabled(True)
    
    def run(self):
        """Cargar movimientos desde la base de datos en un hilo separado"""
        try:
            if not self._is_running:
                logging.info("Thread cancelado antes de iniciar")
                return
            
            logging.info(f"Cargando movimientos (límite: {self.limite})...")
            # Usar el método de postgres_manager que retorna movimientos completos
            rows = self.pg_manager.obtener_movimientos_completos(limite=self.limite)
            
            if self._is_running:  # Verificar si el thread aún debe ejecutarse
                logging.info(f"✅ Thread obtuvo {len(rows)} registros")
                self.movimientos_loaded.emit(rows)
            else:
                logging.info("Thread cancelado antes de emitir datos")
                
        except Exception as e:
            if self._is_running:
                logging.error(f"Error en thread de movimientos: {e}")
                self.error_occurred.emit(str(e))
            else:
                logging.info(f"Error en thread cancelado: {e}")
    
    def stop(self):
        """Detener el thread de forma segura"""
        logging.info("🛑 Señal de parada enviada al thread")
        self._is_running = False


class HistorialMovimientosWindow(QWidget):
    """Widget para ver el historial completo de movimientos de inventario"""
    
    cerrar_solicitado = Signal()
    
    def __init__(self, pg_manager, user_data, parent=None):
        super().__init__(parent)
        self.pg_manager = pg_manager
        self.user_data = user_data
        self.movimientos_data = []
        self.movimientos_filtrados = []
        self.loader_thread = None
        self.pagina_actual = 0  # Para paginación
        self.items_por_pagina = 50
        self.total_movimientos_disponibles = 0
        self._is_visible = True  # Rastrear visibilidad
        
        self.setup_ui()
        self.cargar_movimientos()
    
    def hideEvent(self, event):
        """Evento cuando el widget se oculta (cambio de pestaña)"""
        self._is_visible = False
        self.detener_carga()
        super().hideEvent(event)
    
    def showEvent(self, event):
        """Evento cuando el widget se muestra nuevamente"""
        self._is_visible = True
        super().showEvent(event)
    
    def detener_carga(self):
        """Detener cualquier carga en progreso - Llamado cuando se cambia de ventana"""
        try:
            if self.loader_thread and self.loader_thread.isRunning():
                logging.info("🛑 Deteniendo carga de movimientos (ventana oculta)")
                self.loader_thread.stop()
                self.loader_thread.quit()
                if not self.loader_thread.wait(500):
                    self.loader_thread.terminate()
                    self.loader_thread.wait(500)
        except Exception as e:
            logging.error(f"Error al detener carga: {e}")
    
    def setup_ui(self):
        """Configurar interfaz del historial"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Contenido
        content = QWidget()
        content_layout = create_page_layout("")
        content.setLayout(content_layout)
        
        # Panel de filtros
        filters_panel = self.create_filters_panel()
        content_layout.addWidget(filters_panel)
        
        # Panel para la tabla
        table_panel = self.create_table_panel()
        content_layout.addWidget(table_panel)
        
        # Panel de información y botones (incluye paginación integrada)
        info_buttons_panel = self.create_info_buttons_panel()
        content_layout.addWidget(info_buttons_panel)
        
        layout.addWidget(content)
    
    def create_filters_panel(self):
        """Crear el panel de filtros"""
        filters_panel = ContentPanel()
        filters_layout = QVBoxLayout(filters_panel)
        
        # Primera fila de filtros
        filters_row1 = QHBoxLayout()
        filters_row1.setSpacing(WindowsPhoneTheme.MARGIN_MEDIUM)
        
        # Buscador
        self.search_bar = SearchBar("Buscar por código, nombre o usuario...")
        self.search_bar.connect_search(self.aplicar_filtros)
        filters_row1.addWidget(self.search_bar, stretch=3)
        
        # Filtro por tipo de movimiento
        tipo_container = self.create_tipo_filter()
        filters_row1.addWidget(tipo_container, stretch=1)
        
        filters_layout.addLayout(filters_row1)
        
        # Segunda fila - Rango de fechas
        filters_row2 = self.create_fecha_filters()
        filters_layout.addLayout(filters_row2)
        
        return filters_panel
    
    def create_tipo_filter(self):
        """Crear el filtro por tipo de movimiento"""
        tipo_container = QWidget()
        tipo_layout = QVBoxLayout(tipo_container)
        tipo_layout.setContentsMargins(0, 0, 0, 0)
        tipo_layout.setSpacing(4)
        
        tipo_label = StyledLabel("Tipo:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        tipo_layout.addWidget(tipo_label)
        
        self.tipo_combo = QComboBox()
        self.tipo_combo.addItems([
            "Todos",
            "Entrada",
            "Salida",
            "Ajuste"
        ])
        self.tipo_combo.setMinimumHeight(40)
        self.tipo_combo.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        self.tipo_combo.currentTextChanged.connect(self.aplicar_filtros)
        tipo_layout.addWidget(self.tipo_combo)
        
        return tipo_container
    
    def create_fecha_filters(self):
        """Crear los filtros de fecha"""
        filters_row2 = QHBoxLayout()
        filters_row2.setSpacing(WindowsPhoneTheme.MARGIN_MEDIUM)
        
        # Fecha inicio
        fecha_inicio_container = QWidget()
        fecha_inicio_layout = QVBoxLayout(fecha_inicio_container)
        fecha_inicio_layout.setContentsMargins(0, 0, 0, 0)
        fecha_inicio_layout.setSpacing(4)
        
        fecha_inicio_label = StyledLabel("Desde:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        fecha_inicio_layout.addWidget(fecha_inicio_label)
        
        self.fecha_inicio = QDateEdit()
        self.fecha_inicio.setDate(QDate.currentDate().addMonths(-1))
        self.fecha_inicio.setCalendarPopup(True)
        aplicar_estilo_fecha(self.fecha_inicio)
        self.fecha_inicio.setMinimumHeight(40)
        self.fecha_inicio.dateChanged.connect(self.aplicar_filtros)
        fecha_inicio_layout.addWidget(self.fecha_inicio)
        
        filters_row2.addWidget(fecha_inicio_container, stretch=1)
        
        # Fecha fin
        fecha_fin_container = QWidget()
        fecha_fin_layout = QVBoxLayout(fecha_fin_container)
        fecha_fin_layout.setContentsMargins(0, 0, 0, 0)
        fecha_fin_layout.setSpacing(4)
        
        fecha_fin_label = StyledLabel("Hasta:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        fecha_fin_layout.addWidget(fecha_fin_label)
        
        self.fecha_fin = QDateEdit()
        self.fecha_fin.setDate(QDate.currentDate())
        self.fecha_fin.setCalendarPopup(True)
        aplicar_estilo_fecha(self.fecha_fin)
        self.fecha_fin.setMinimumHeight(40)
        self.fecha_fin.dateChanged.connect(self.aplicar_filtros)
        fecha_fin_layout.addWidget(self.fecha_fin)
        
        filters_row2.addWidget(fecha_fin_container, stretch=1)
        
        # Botón limpiar filtros
        btn_limpiar = QPushButton("Limpiar Filtros")
        btn_limpiar.setMinimumHeight(40)
        btn_limpiar.setMinimumWidth(120)
        btn_limpiar.setObjectName("tileButton")
        btn_limpiar.setProperty("tileColor", WindowsPhoneTheme.TILE_ORANGE)
        btn_limpiar.clicked.connect(self.limpiar_filtros)
        filters_row2.addWidget(btn_limpiar, alignment=Qt.AlignBottom)
        
        return filters_row2
    
    def create_table_panel(self):
        """Crear el panel con la tabla de movimientos"""
        table_panel = ContentPanel()
        table_layout = QVBoxLayout(table_panel)
        table_layout.setContentsMargins(0, 0, 0, 0)
        
        # Tabla de movimientos
        self.movimientos_table = QTableWidget()
        self.movimientos_table.setColumnCount(9)
        self.movimientos_table.setHorizontalHeaderLabels([
            "Fecha", "Tipo", "Código", "Producto", "Cantidad", 
            "Stock Ant.", "Stock Nuevo", "Motivo", "Usuario"
        ])
        
        # Configurar header
        header = self.movimientos_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.Stretch)
        header.setSectionResizeMode(8, QHeaderView.ResizeToContents)
        
        # Estilo de la tabla
        self.movimientos_table.setAlternatingRowColors(True)
        self.movimientos_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.movimientos_table.setSelectionMode(QTableWidget.SingleSelection)
        self.movimientos_table.verticalHeader().setVisible(False)
        self.movimientos_table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        table_layout.addWidget(self.movimientos_table)
        return table_panel
    
    def create_pagination_panel(self):
        """Crear el panel de paginación INTEGRADO en la fila inferior"""
        # Nota: Este método ya no crea un panel separado
        # Los botones se integran en create_info_buttons_panel
        pass  # Los botones se crean en create_info_buttons_panel
    
    def actualizar_pagination_buttons(self):
        """Actualizar estado de botones de paginación e info"""
        total_paginas = (len(self.movimientos_filtrados) + self.items_por_pagina - 1) // self.items_por_pagina
        
        self.btn_pagina_anterior.setEnabled(self.pagina_actual > 0)
        self.btn_proxima_pagina.setEnabled(self.pagina_actual < total_paginas - 1)
        
        inicio = self.pagina_actual * self.items_por_pagina + 1
        fin = min((self.pagina_actual + 1) * self.items_por_pagina, len(self.movimientos_filtrados))
        total_filtrados = len(self.movimientos_filtrados)
        total_disponibles = len(self.movimientos_data)  # Total de todos los movimientos cargados
        
        # Actualizar label con información completa de paginación
        if total_filtrados > 0:
            if total_paginas > 1:
                # Mostrar info completa: qué se muestra + página actual + totales
                self.info_label.setText(
                    f"Página {self.pagina_actual + 1}/{total_paginas} | Mostrando {inicio}-{fin} de {total_filtrados} | Total disponible: {total_disponibles}"
                )
            else:
                # Solo una página pero mostrar total disponible
                self.info_label.setText(f"Total: {total_filtrados} movimientos (de {total_disponibles} disponibles)")
        else:
            self.info_label.setText("No hay registros")
    
    def pagina_anterior(self):
        """Ir a la página anterior"""
        if self.pagina_actual > 0:
            self.pagina_actual -= 1
            self.mostrar_pagina_actual()
    
    def proxima_pagina(self):
        """Ir a la siguiente página"""
        total_paginas = (len(self.movimientos_filtrados) + self.items_por_pagina - 1) // self.items_por_pagina
        if self.pagina_actual < total_paginas - 1:
            self.pagina_actual += 1
            self.mostrar_pagina_actual()
    
    def mostrar_pagina_actual(self):
        """Mostrar la página actual de movimientos"""
        inicio = self.pagina_actual * self.items_por_pagina
        fin = inicio + self.items_por_pagina
        movimientos_pagina = self.movimientos_filtrados[inicio:fin]
        self.mostrar_movimientos(movimientos_pagina, mostrar_paginacion=True)
    
    def create_info_buttons_panel(self):
        """Crear el panel de información y botones con paginación integrada"""
        info_buttons_panel = ContentPanel()
        info_buttons_layout = QHBoxLayout(info_buttons_panel)
        info_buttons_layout.setSpacing(8)
        
        # Etiqueta de información (con paginación integrada)
        self.info_label = StyledLabel("", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        info_buttons_layout.addWidget(self.info_label, stretch=1)
        
        # Separador visual
        separator = QFrame()
        separator.setFrameShape(QFrame.VLine)
        separator.setFrameShadow(QFrame.Sunken)
        separator.setMaximumWidth(1)
        info_buttons_layout.addWidget(separator)
        
        # Botones de paginación COMPACTOS con diseño consistente
        # Botón anterior: Icono (◀) + Texto
        self.btn_pagina_anterior = CompactNavButton(
            "Anterior", 
            "fa5s.chevron-left",
            WindowsPhoneTheme.TILE_BLUE,
            icon_position="left"  # Icono a la izquierda
        )
        self.btn_pagina_anterior.setToolTip("Página anterior")
        self.btn_pagina_anterior.clicked.connect(self.pagina_anterior)
        info_buttons_layout.addWidget(self.btn_pagina_anterior)
        
        # Botón siguiente: Texto + Icono (▶)
        self.btn_proxima_pagina = CompactNavButton(
            "Siguiente",
            "fa5s.chevron-right",
            WindowsPhoneTheme.TILE_BLUE,
            icon_position="right"  # Icono a la derecha
        )
        self.btn_proxima_pagina.setToolTip("Página siguiente")
        self.btn_proxima_pagina.clicked.connect(self.proxima_pagina)
        info_buttons_layout.addWidget(self.btn_proxima_pagina)
        
        # Separador visual
        separator2 = QFrame()
        separator2.setFrameShape(QFrame.VLine)
        separator2.setFrameShadow(QFrame.Sunken)
        separator2.setMaximumWidth(1)
        info_buttons_layout.addWidget(separator2)
        
        # Botones de acción principales
        btn_exportar = TileButton("Exportar Excel", "fa5s.file-excel", WindowsPhoneTheme.TILE_GREEN)
        btn_exportar.clicked.connect(self.exportar_excel)
        btn_exportar.setMaximumWidth(180)
        info_buttons_layout.addWidget(btn_exportar)
        
        btn_actualizar = TileButton("Actualizar", "fa5s.sync", WindowsPhoneTheme.TILE_BLUE)
        btn_actualizar.clicked.connect(self.cargar_movimientos)
        btn_actualizar.setMaximumWidth(150)
        info_buttons_layout.addWidget(btn_actualizar)
        
        btn_volver = TileButton("Volver", "fa5s.arrow-left", WindowsPhoneTheme.TILE_RED)
        btn_volver.clicked.connect(self.cerrar_solicitado.emit)
        btn_volver.setMaximumWidth(120)
        info_buttons_layout.addWidget(btn_volver)
        
        return info_buttons_panel
    
    def cargar_movimientos(self):
        """Cargar todos los movimientos desde la base de datos de forma asíncrona"""
        try:
            # Si no estamos visibles, no cargar
            if not self._is_visible:
                return
                
            # Mostrar indicador de carga
            self.info_label.setText("Cargando movimientos...")
            self.movimientos_table.setRowCount(0)
            self.pagina_actual = 0  # Resetear paginación
            
            # Detener hilo anterior si existe
            if self.loader_thread and self.loader_thread.isRunning():
                self.loader_thread.stop()  # Señalizar al thread que se detenga
                self.loader_thread.quit()
                if not self.loader_thread.wait(1000):  # Esperar hasta 1 segundo
                    self.loader_thread.terminate()
                    self.loader_thread.wait()
            
            # Crear y ejecutar hilo de carga (cargar 500 para tener más datos disponibles)
            self.loader_thread = MovimientosLoaderThread(self.pg_manager, limite=500)
            self.loader_thread.movimientos_loaded.connect(self.procesar_datos_movimientos, Qt.QueuedConnection)
            self.loader_thread.error_occurred.connect(self.mostrar_error_carga, Qt.QueuedConnection)
            self.loader_thread.finished.connect(self.on_thread_finished, Qt.QueuedConnection)
            logging.info("Iniciando carga de movimientos...")
            self.loader_thread.start()
            
        except Exception as e:
            logging.error(f"Error iniciando carga de movimientos: {e}")
            show_error_dialog(
                self,
                "Error al cargar",
                "No se pudieron cargar los movimientos",
                detail=str(e)
            )
    
    def on_thread_finished(self):
        """Callback cuando el thread termina"""
        pass
    
    def procesar_datos_movimientos(self, rows):
        """Procesar los datos de movimientos cargados desde la base de datos"""
        try:
            # Verificar que la ventana aún está visible
            if not self._is_visible:
                logging.info("Descartando datos - ventana no está visible")
                return
            
            self.movimientos_data = []
            
            for row in rows:
                self.movimientos_data.append({
                    'id_movimiento': row['id_movimiento'],
                    'fecha': row['fecha'],
                    'tipo_movimiento': row['tipo_movimiento'],
                    'codigo_interno': row['codigo_interno'],
                    'tipo_producto': row['tipo_producto'],
                    'cantidad': row['cantidad'],
                    'stock_anterior': row['stock_anterior'],
                    'stock_nuevo': row['stock_nuevo'],
                    'motivo': row['motivo'] or '',
                    'id_usuario': row['id_usuario'],
                    'id_venta': row['id_venta'],
                    'nombre_producto': row['nombre_producto'],
                    'nombre_usuario': row['nombre_usuario'] or 'Usuario desconocido'
                })
            
            self.aplicar_filtros()
            logging.info(f"✅ Obtuvieron {len(self.movimientos_data)} movimientos completos")
            
        except Exception as e:
            logging.error(f"Error procesando datos de movimientos: {e}")
            if self._is_visible:
                show_error_dialog(
                    self,
                    "Error al procesar datos",
                    "No se pudieron procesar los datos de movimientos",
                    detail=str(e)
                )
    
    def mostrar_error_carga(self, error_msg):
        """Mostrar mensaje de error al cargar movimientos"""
        # Verificar que la ventana aún está visible
        if not self._is_visible:
            logging.info("Error descartado - ventana no está visible")
            return
            
        logging.error(f"Error cargando movimientos: {error_msg}")
        show_error_dialog(
            self,
            "Error al cargar",
            "No se pudieron cargar los movimientos",
            detail=error_msg
        )
        self.info_label.setText("Error al cargar movimientos")
    
    def aplicar_filtros(self):
        """Aplicar todos los filtros activos"""
        try:
            # Obtener criterios de filtro
            texto_busqueda = self.search_bar.text().strip().lower()
            tipo_seleccionado = self.tipo_combo.currentText()
            fecha_desde = self.fecha_inicio.date().toPython()
            fecha_hasta = self.fecha_fin.date().toPython()
            
            # Filtrar datos
            self.movimientos_filtrados = []
            for mov in self.movimientos_data:
                # Filtro de texto
                if texto_busqueda:
                    if not any([
                        texto_busqueda in str(mov['codigo_interno']).lower(),
                        texto_busqueda in str(mov['nombre_producto']).lower(),
                        texto_busqueda in str(mov['nombre_usuario']).lower(),
                        texto_busqueda in (mov['motivo'] or '').lower()
                    ]):
                        continue
                
                # Filtro de tipo
                if tipo_seleccionado != "Todos":
                    if mov['tipo_movimiento'].lower() != tipo_seleccionado.lower():
                        continue
                
                # Filtro de fecha - Convertir fecha_mov a date object si es datetime
                fecha_mov = mov['fecha']
                if isinstance(fecha_mov, datetime):
                    fecha_mov = fecha_mov.date()
                elif isinstance(fecha_mov, str):
                    try:
                        # Si es string, intentar parsearlo
                        fecha_mov = datetime.fromisoformat(fecha_mov.replace('Z', '+00:00')).date()
                    except:
                        continue
                
                # Ahora comparar date con date
                if not (fecha_desde <= fecha_mov <= fecha_hasta):
                    continue
                
                self.movimientos_filtrados.append(mov)
            
            # Resetear paginación cuando se aplican nuevos filtros
            self.pagina_actual = 0
            self.mostrar_pagina_actual()
            
        except Exception as e:
            logging.error(f"Error aplicando filtros: {e}")
            self.movimientos_filtrados = self.movimientos_data
            self.pagina_actual = 0
            self.mostrar_pagina_actual()
    
    def mostrar_movimientos(self, movimientos, mostrar_paginacion=False):
        """Mostrar movimientos en la tabla"""
        try:
            self.movimientos_table.setRowCount(0)
            
            for mov in movimientos:
                row = self.movimientos_table.rowCount()
                self.movimientos_table.insertRow(row)
                
                # Fecha
                fecha_str = mov['fecha'].strftime("%d/%m/%Y %H:%M") if isinstance(mov['fecha'], datetime) else str(mov['fecha'])
                item_fecha = QTableWidgetItem(fecha_str)
                item_fecha.setTextAlignment(Qt.AlignCenter)
                self.movimientos_table.setItem(row, 0, item_fecha)
                
                # Tipo de movimiento
                tipo = mov['tipo_movimiento'].capitalize()
                item_tipo = QTableWidgetItem(tipo)
                item_tipo.setTextAlignment(Qt.AlignCenter)
                
                # Color según tipo
                if mov['tipo_movimiento'].lower() == 'entrada':
                    item_tipo.setForeground(Qt.darkGreen)
                elif mov['tipo_movimiento'].lower() == 'salida':
                    item_tipo.setForeground(Qt.darkRed)
                else:
                    item_tipo.setForeground(Qt.darkBlue)
                
                self.movimientos_table.setItem(row, 1, item_tipo)
                
                # Código interno
                self.movimientos_table.setItem(row, 2, QTableWidgetItem(str(mov['codigo_interno'])))
                
                # Nombre producto
                self.movimientos_table.setItem(row, 3, QTableWidgetItem(str(mov['nombre_producto'])))
                
                # Cantidad
                cantidad = mov['cantidad']
                item_cantidad = QTableWidgetItem(str(cantidad))
                item_cantidad.setTextAlignment(Qt.AlignCenter)
                if cantidad > 0:
                    item_cantidad.setForeground(Qt.darkGreen)
                else:
                    item_cantidad.setForeground(Qt.darkRed)
                self.movimientos_table.setItem(row, 4, item_cantidad)
                
                # Stock anterior
                item_stock_ant = QTableWidgetItem(str(mov['stock_anterior']))
                item_stock_ant.setTextAlignment(Qt.AlignCenter)
                self.movimientos_table.setItem(row, 5, item_stock_ant)
                
                # Stock nuevo
                item_stock_nuevo = QTableWidgetItem(str(mov['stock_nuevo']))
                item_stock_nuevo.setTextAlignment(Qt.AlignCenter)
                self.movimientos_table.setItem(row, 6, item_stock_nuevo)
                
                # Motivo
                self.movimientos_table.setItem(row, 7, QTableWidgetItem(str(mov['motivo'] or '')))
                
                # Usuario
                item_usuario = QTableWidgetItem(str(mov['nombre_usuario']))
                item_usuario.setTextAlignment(Qt.AlignCenter)
                self.movimientos_table.setItem(row, 8, item_usuario)
            
            # Actualizar información y paginación
            if mostrar_paginacion:
                self.actualizar_pagination_buttons()
            else:
                # No mostrar paginación en label, solo contar
                total_movimientos = len(movimientos)
                # (El info_label se actualiza en actualizar_pagination_buttons cuando se llama desde mostrar_pagina_actual)
            
            logging.info(f"Mostrando {len(movimientos)} movimientos en tabla")
            
        except Exception as e:
            logging.error(f"Error mostrando movimientos: {e}")
            show_error_dialog(
                self,
                "Error de visualización",
                "No se pudieron mostrar los movimientos",
                detail=str(e)
            )
    
    def limpiar_filtros(self):
        """Limpiar todos los filtros y mostrar todo"""
        self.search_bar.clear()
        self.tipo_combo.setCurrentIndex(0)
        self.fecha_inicio.setDate(QDate.currentDate().addMonths(-1))
        self.fecha_fin.setDate(QDate.currentDate())
        self.aplicar_filtros()
    
    def exportar_excel(self):
        """Exportar movimientos filtrados a Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Movimientos Inventario"
            
            # Encabezados
            headers = [
                "Fecha", "Tipo", "Código", "Producto", "Cantidad", 
                "Stock Anterior", "Stock Nuevo", "Motivo", "Usuario", "ID Venta"
            ]
            
            # Estilo de encabezado
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Obtener movimientos filtrados actuales
            texto_busqueda = self.search_bar.text().strip().lower()
            tipo_seleccionado = self.tipo_combo.currentText()
            fecha_desde = self.fecha_inicio.date().toPython()
            fecha_hasta = self.fecha_fin.date().toPython()
            
            movimientos_exportar = []
            for mov in self.movimientos_data:
                if texto_busqueda:
                    if not any([
                        texto_busqueda in str(mov['codigo_interno']).lower(),
                        texto_busqueda in str(mov['nombre_producto']).lower(),
                        texto_busqueda in str(mov['nombre_usuario']).lower(),
                        texto_busqueda in (mov['motivo'] or '').lower()
                    ]):
                        continue
                
                if tipo_seleccionado != "Todos":
                    if mov['tipo_movimiento'].lower() != tipo_seleccionado.lower():
                        continue
                
                # Convertir fecha_mov a date object si es datetime
                fecha_mov = mov['fecha']
                if isinstance(fecha_mov, datetime):
                    fecha_mov = fecha_mov.date()
                elif isinstance(fecha_mov, str):
                    try:
                        # Si es string, intentar parsearlo
                        fecha_mov = datetime.fromisoformat(fecha_mov.replace('Z', '+00:00')).date()
                    except:
                        continue
                
                # Ahora comparar date con date
                if not (fecha_desde <= fecha_mov <= fecha_hasta):
                    continue
                
                movimientos_exportar.append(mov)
            
            # Datos
            for row_idx, mov in enumerate(movimientos_exportar, start=2):
                fecha_str = mov['fecha'].strftime("%d/%m/%Y %H:%M") if isinstance(mov['fecha'], datetime) else str(mov['fecha'])
                
                ws.cell(row=row_idx, column=1, value=fecha_str)
                ws.cell(row=row_idx, column=2, value=mov['tipo_movimiento'].capitalize())
                ws.cell(row=row_idx, column=3, value=mov['codigo_interno'])
                ws.cell(row=row_idx, column=4, value=mov['nombre_producto'])
                ws.cell(row=row_idx, column=5, value=mov['cantidad'])
                ws.cell(row=row_idx, column=6, value=mov['stock_anterior'])
                ws.cell(row=row_idx, column=7, value=mov['stock_nuevo'])
                ws.cell(row=row_idx, column=8, value=mov['motivo'])
                ws.cell(row=row_idx, column=9, value=mov['nombre_usuario'])
                ws.cell(row=row_idx, column=10, value=mov['id_venta'] if mov['id_venta'] else '')
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 30
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 10
            
            # Guardar archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"movimientos_inventario_{timestamp}.xlsx"
            wb.save(filename)
            
            show_info_dialog(
                self,
                "Exportación exitosa",
                f"Archivo generado: {filename}\n\nMovimientos exportados: {len(movimientos_exportar)}"
            )
            
            logging.info(f"Reporte de movimientos exportado: {filename}")
            
        except ImportError:
            show_error_dialog(
                self,
                "Módulo no disponible",
                "No se puede exportar a Excel. El módulo openpyxl no está instalado."
            )
        except Exception as e:
            logging.error(f"Error exportando a Excel: {e}")
            show_error_dialog(
                self,
                "Error de exportación",
                "No se pudo generar el archivo Excel",
                detail=str(e)
            )
    
    def closeEvent(self, event):
        """Evento al cerrar la ventana - Limpiar threads correctamente"""
        try:
            self._is_visible = False
            # Detener y limpiar el thread de forma segura
            if self.loader_thread:
                if self.loader_thread.isRunning():
                    logging.info("🛑 Deteniendo thread durante closeEvent")
                    self.loader_thread.stop()  # Señalizar al thread que se detenga
                    # Desconectar todas las señales
                    try:
                        self.loader_thread.movimientos_loaded.disconnect()
                    except:
                        pass
                    try:
                        self.loader_thread.error_occurred.disconnect()
                    except:
                        pass
                    try:
                        self.loader_thread.finished.disconnect()
                    except:
                        pass
                    
                    self.loader_thread.quit()
                    if not self.loader_thread.wait(1500):  # Esperar hasta 1.5 segundos
                        self.loader_thread.terminate()
                        self.loader_thread.wait()
                self.loader_thread = None
        except Exception as e:
            logging.error(f"Error al cerrar ventana de movimientos: {e}")
        finally:
            event.accept()
    
    def __del__(self):
        """Destructor para limpiar el thread"""
        try:
            if hasattr(self, '_is_visible'):
                self._is_visible = False
            if hasattr(self, 'loader_thread') and self.loader_thread:
                if self.loader_thread.isRunning():
                    self.loader_thread.stop()
                    self.loader_thread.quit()
                    self.loader_thread.wait(1000)
        except (RuntimeError, AttributeError):
            # El thread ya fue eliminado por C++
            pass